import scrapy
import openpyxl
import sys
from datetime import datetime

class InstitutionenSpider(scrapy.Spider):
    name = 'institutionen'
    allowed_domains = ['www.heiminfo.ch']
    start_urls = []
    elements = []

    #CONSTRUCTOR TO INITIALIZE START URLS LIST AND SEE WHICH ELEMENTS TO GET
    def __init__(self):
        input_file = open("all_urls.txt","r")

        all_lines = input_file.readlines()

        for one_line in all_lines:
            if one_line.strip() == ("" or "\n"):
                continue
            temp = one_line.strip()
            temp = one_line.strip('\n')
            self.start_urls.append(temp)

        input_file.close()

        self.elements = self.elements_to_get()

    #THIS IS USED TO GET WOHN
    def get_wohn(self,response):
        Wohn1 = response.xpath("normalize-space(//div[@class='rooms']/ul/li[1])").get()
        if Wohn1 is not None:
            Wohn1 = Wohn1.replace('\t','')
            Wohn1 = Wohn1.replace('\n','')
            #Wohn1 = Wohn1 + " " + response.xpath("normalize-space(//div[@class='rooms']/ul/li[1]/small/text())").get()

        Wohn2 = response.xpath("normalize-space(//div[@class='rooms']/ul/li[2])").get()
        if Wohn2 is not None:
            Wohn2 = Wohn2.replace('\t','')
            Wohn2 = Wohn2.replace('\n','')
            #Wohn2 = Wohn2 + " " + response.xpath("normalize-space(//div[@class='rooms']/ul/li[2]/small/text())").get()

        Wohn3 = response.xpath("normalize-space(//div[@class='rooms']/ul/li[3])").get()
        if Wohn3 is not None:
            Wohn3 = Wohn3.replace('\t','')
            Wohn3 = Wohn3.replace('\n','')
            #Wohn3 = Wohn3 + " " + response.xpath("normalize-space(//div[@class='rooms']/ul/li[3]/small/text())").get()

        Wohn4 = response.xpath("normalize-space(//div[@class='rooms']/ul/li[4])").get()
        if Wohn4 is not None:
            Wohn4 = Wohn4.replace('\t','')
            Wohn4 = Wohn4.replace('\n','')
            #Wohn4 = Wohn4 + " " + response.xpath("normalize-space(//div[@class='rooms']/ul/li[4]/small/text())").get()

        Wohn = []
        Wohn.append(Wohn1)
        Wohn.append(Wohn2)
        Wohn.append(Wohn3)
        Wohn.append(Wohn4)
        return Wohn

    #THIS IS USED TO GET VARIOUS ADDRESS FIELDS
    def get_addresses(self,response):
        empty = []

        #PO BOX NO IS INCLUDED
        if response.xpath("(//p[@class='location']/span[6])[2]/text()").get() is not None:
            strabe = response.xpath("normalize-space((//p[@class='location']/span[1])[2]/text())").get()
            strabe = strabe + " " + response.xpath("normalize-space((//p[@class='location']/span[2])[2]/text())").get()
            strabe = strabe + " " + response.xpath("normalize-space((//p[@class='location']/span[3])[2]/text())").get()
            plz = response.xpath("normalize-space((//p[@class='location']/span[4])[2]/text())").get()
            ort = response.xpath("normalize-space((//p[@class='location']/span[5])[2]/text())").get()
        elif response.xpath("(//p[@class='location']/span[5])[2]/text()").get() is not None:
            strabe = response.xpath("normalize-space((//p[@class='location']/span[1])[2]/text())").get()
            strabe = strabe + " " + response.xpath("normalize-space((//p[@class='location']/span[2])[2]/text())").get()
            plz = response.xpath("normalize-space((//p[@class='location']/span[3])[2]/text())").get()
            ort = response.xpath("normalize-space((//p[@class='location']/span[4])[2]/text())").get()
        else:
            strabe = response.xpath("normalize-space((//p[@class='location']/span[1])[2]/text())").get()
            plz = response.xpath("normalize-space((//p[@class='location']/span[2])[2]/text())").get()
            ort = response.xpath("normalize-space((//p[@class='location']/span[3])[2]/text())").get()

        strabe = strabe.replace('\t','').replace('\n','').strip()
        plz = plz.replace('\t','').replace('\n','').strip()
        ort = ort.replace('\t','').replace('\n','').strip()

        empty.append(strabe)
        empty.append(plz)
        empty.append(ort)

        return empty

    #THIS IS USED TO GET NAME
    def get_name(self,response):
        name = response.xpath("normalize-space(//div[@class='head']/h2/text())").get()
        if name is not None:
            name = name.replace('\t','').replace('\n','').strip()
        return name

    #THIS IS USED TO GET C/O
    def get_co(self,response):
        co = response.xpath("normalize-space((//div[@class='head']/p)[2]/text())").get()
        if co is not None:
            co = co.replace('\t','').replace('\n','').strip()
        return co

    #THIS IS USED TO GET TELEPHONE NO
    def get_tel_no(self,response):
        number = response.xpath("(//p[@class='phone'])[position()=last()]/text()").getall()
        if number is not None:
            if len(number) > 1:
                n1 = number[1]
            else:
                n1 = number[0]
            n1 = n1.replace('\t','').replace('\n','').strip()
            return n1
        else:
            return None

    #THIS IS USED TO GET WEBSITE
    def get_website(self,response):
        site = response.xpath("normalize-space(//div[@class='website-and-map-container']/a[1]/@href)").get()
        if site is not None:
            site = site.replace('\t','').replace('\n','').strip()
        return site

    #THIS IS THE UI FUNCTION
    def user_interface(self):
        print("\n##################################################################################")
        print("Hello, Please Select One of The Following Options:-\n")
        print("1. The File 'Parameter.xlsx' is Present in The Same Directory as This Program")
        print("2. The File 'Parameter.xlsx' is Present in Some Other Directory")
        print("##################################################################################")

        while(1):
            choice = input("\n\tPlease Enter 1 or 2 to Select Your Option : ")
            if choice != "1" and choice != "2":
                print("\n\t####You have Entered a Wrong Option, Please Try Again####")
                continue
            else:
                break
                
        if choice == "1":
            path = "Parameter.xlsx"

        if choice == "2": 
            while(1):
                path = input("\n\t\tPlease Enter The Full Path of The Parameter File : ")
                try:
                    wb = openpyxl.load_workbook(path)
                    wb.close()
                    break
                except FileNotFoundError:
                    print("\n\t####Error, You Have Entered The Wrong Path####")
                    continue    
        return path

    #THIS IS USED TO SEE WHICH ELEMENTS SHOULD WE GET
    def elements_to_get(self):
        path = self.user_interface()
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb['Para']

        data_list = []
        val_list = []
        list_of_list = []

        print("\n")
        print(f"EXCEL FILE {path} OPENED")
        print("SHEET 'Para' OPENED")
        print("\n")

        website_name = ws['A3'].value     
        print('CELL A3 ACCESSED',website_name)

        #CHECKING APPROPIATE VALUE OF THAT CELL
        if website_name is None:
            sys.exit("ERROR ! CELL A3 IS EMPTY")
        elif "heiminfo.ch" not in website_name.lower():
            sys.exit("ERROR ! CELL A3 DOES NOT CONTAIN 'heiminfo.ch'")

        structure = ws['B3'].value
        print('CELL B3 ACCESSED',structure)

        if structure.strip() not in wb.sheetnames:
            sys.exit(f"ERROR ! THERE IS NO SHEET NAMED '{structure.strip()}' IN THIS EXCEL FILE")

        #OPENING THE STRUCTURE SHEET
        ws = wb[structure.strip()]

        #GETTING THE DATA STRUCTURE FROM THAT SHEET
        for index, row in enumerate(ws.iter_rows()):

            print(index+1, 'Row Accessed')

            if row[0].value is None:
                continue
            elif row[0].value.strip() == "":
                continue

            #GETTING THE VALUES
            data = row[0].value
            val = row[1].value

            #SAVING THE VALUES
            data_list.append(data)
            val_list.append(val)

        #APPENDING TO LIST OF LIST
        list_of_list.append(data_list)
        list_of_list.append(val_list)

        return list_of_list

    #MAIN PARSE FUNCTION
    def parse(self, response):

        addresses = self.get_addresses(response)
        wohn = self.get_wohn(response)

        item_dict = {}
        
        for index, val in enumerate(self.elements[0]):

            #PARAMETER FILE COMPONENTS
            if "Daten-Prioritaet" in val:
                item_dict[f'{val}'] = self.elements[1][index]
            elif "Datenart" in val:
                item_dict[f'{val}'] = self.elements[1][index]
            elif "Branche" in val:
                item_dict[f'{val}'] = self.elements[1][index]
            elif "Betriebstyp" in val:
                item_dict[f'{val}'] = self.elements[1][index]
            elif "NOGA" in val:
                item_dict[f'{val}'] = self.elements[1][index]
            elif "Erfassungsdatum" in val:
                item_dict[f'{val}'] = datetime.today().strftime('%d.%m.%Y')
            elif "Operator" in val:
                item_dict[f'{val}'] = self.elements[1][index]
            elif "Wie" in val:
                item_dict[f'{val}'] = self.elements[1][index]

            #SCRAPED COMPONENTS
            elif "Firmenname" in val:
                item_dict[f'{val}'] = self.get_name(response)
            elif "c/o" in val:
                item_dict[f'{val}'] = self.get_co(response)
            elif "Straße" in val:
                item_dict[f'{val}'] = addresses[0]
            elif "PLZ" in val:
                item_dict[f'{val}'] = addresses[1]
            elif "Ort" in val:
                item_dict[f'{val}'] = addresses[2]
            elif "Telefonnummer" in val:
                item_dict[f'{val}'] = self.get_tel_no(response)
            elif "Website" in val:
                item_dict[f'{val}'] = self.get_website(response)
            elif "Wohn1" in val:
                item_dict[f'{val}'] = wohn[0]
            elif "Wohn2" in val:
                item_dict[f'{val}'] = wohn[1]
            elif "Wohn3" in val:
                item_dict[f'{val}'] = wohn[2]
            elif "Wohn4" in val:
                item_dict[f'{val}'] = wohn[3]

            #REST DATA
            else:
                item_dict[f'{val}'] = ''

        yield item_dict